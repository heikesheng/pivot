import pandas as pd
import os
from openpyxl import load_workbook

##此脚本实现汇率转换，结合汇率表转化每个国家产出表的value
# 定义文件夹路径
export_dir = '/Users/zxt/Desktop/ioput'  # 出口Excel表的目录
exchange_file_path = '/Users/zxt/Desktop/exchange/exchange.xlsx'  # 汇率Excel表的路径

# 读取汇率表
df_exchange = pd.read_excel(exchange_file_path, sheet_name='Sheet1')

# 将汇率表从宽格式转换为长格式
df_exchange_long = df_exchange.melt(id_vars=['Country'], var_name='Year', value_name='Exchange_Rate')

# 确保Year列为整数类型，以便与出口表的Year匹配
df_exchange_long['Year'] = df_exchange_long['Year'].astype(int)

# 遍历出口文件目录
for export_file in os.listdir(export_dir):
    if export_file.endswith('.xlsx'):
        # 读取出口表的第二个sheet
        export_file_path = os.path.join(export_dir, export_file)
        df_export = pd.read_excel(export_file_path, sheet_name=1, engine='openpyxl')  # 读取第二个Sheet
        
        # 确保出口表中有Country, Year, Value列
        if not all(col in df_export.columns for col in ['Country or Area', 'Year', 'Value']):
            print(f"文件 {export_file} 缺少必要列，跳过...")
            continue

        # 将Value列转换为数字格式，移除非数字字符
        df_export['Value'] = pd.to_numeric(df_export['Value'].astype(str).str.replace(',', '').str.strip(), errors='coerce')

        # 检查是否有无法转换为数字的值
        if df_export['Value'].isnull().any():
            print(f"文件 {export_file} 中存在无法转换为数字的 'Value' 数据，结果可能不准确。")

        # 合并出口表和汇率表
        df_merged = df_export.merge(df_exchange_long, left_on=['Country or Area', 'Year'], right_on=['Country', 'Year'], how='left')
        
        # 确保 Exchange_Rate 列为浮点数类型（已在转换汇率表时处理）
        df_merged['Exchange_Rate'] = pd.to_numeric(df_merged['Exchange_Rate'], errors='coerce')

        # 检查合并后是否有汇率缺失
        if df_merged['Exchange_Rate'].isnull().any():
            print(f"文件 {export_file} 中有部分行的汇率缺失，结果可能不准确。")
        
        # 计算Value_in_USD
        df_merged['Value_in_USD'] = df_merged['Value'] / df_merged['Exchange_Rate']

        # 使用openpyxl加载原始Excel文件
        workbook = load_workbook(export_file_path)
        sheet = workbook.worksheets[1]  # 获取第二个sheet
        
        # 将计算结果写回到原Excel文件的第二个sheet中，添加新列Value_in_USD
        # 首先检查是否已经存在Value_in_USD列，如果存在则更新，否则追加
        headers = [cell.value for cell in sheet[1]]  # 获取标题行的所有值
        if 'Value_in_USD' not in headers:
            sheet.cell(row=1, column=len(headers) + 1, value='Value_in_USD')  # 添加标题
        
        # 写入Value_in_USD数据到每一行
        for i, value in enumerate(df_merged['Value_in_USD'], start=2):  # 从第二行开始写，假设第一行是标题
            sheet.cell(row=i, column=len(headers) + 1, value=value)  # 在最后一列添加新数据
        
        # 保存Excel文件
        workbook.save(export_file_path)
        print(f"已处理并保存文件：{export_file_path}")
